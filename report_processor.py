import datetime
from pathlib import Path
import base64
from io import BytesIO
from collections import defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image as XLImage
# ---------------------------------------------------------------------------
# Embedded BSNL logo (stored permanently inside this Python file)
# No external BSNL logo.png file is required.
# ---------------------------------------------------------------------------
BSNL_LOGO_B64 = """iVBORw0KGgoAAAANSUhEUgAAAUAAAACQCAMAAABOB0IDAAAC/VBMVEVHcEz7ax75ayH7ax77ax4HA437ax4HA40HA435ayH7ax77ax77ax4Eazj7ax77ax4HA40HA40HA437ax4HA43yaywHA43yayz7ax77ax77ax4HA40EazgEazgHA40HA43yayzyaywHA43yayzyaywHA40EazgHA43yayzyayz///8EazgHA43yayz///8EazgEazgEazgEazjyaywEazjyayzyayz////zbC7///////////8EazgEazgEazj///////////////////////8Eazj///+HbzcHA43yaywEazj////7ax7ybS/xbzLxcDPwdDrwdTvxbjHvekPugU7ybC3ugE3vfUjug1Hwdj3xcTXybC7xcTbuglDxcjbweEDvfknthlbwdz/thVXve0XweEHvfEfsjGDweULxczjuf0vth1jthFTsjmL0poHuhFPsil31pIDzp4PtiFryrY3uf0ztiVvybjD0o3zzqIbzdDnwdj7rkGXwcznsi1/1oHjrkWfyq4rsi17ve0b0fkbxcjjyr5HyspT2mGz2oXryuZ/1nXPsilz22s7tiFnwdz713NHys5byro/ysJLskmnzrIzrlGz1n3fytZrzxrLsj2T228/11sj1nnXyqojqlm/23tPywKrzvaTzy7nzsZP0zr70ybX12Mvzu6L0oXvyuJ31m3H0pH/veUPyxK/th1n2mW7yqYjwwq3ytJj0wqzzr5D51cXzekLyv6f1zbzql3Hytpvz0MH10sLufkrxyLbyuqH++vjyq4n3q4f1xa/01MX86eD97eb++PXrmXT85Nrqmnbwi1z0tZj1mnD32szwvab60r/98evqmHP5yLH72sr73tD4wKYTbjv6zbnpm3j74dX0kGL+9fH5tZYybTn0iFb4uZv3sZBhimExeUucnnz4vKDhbjFhbDXEcDbBqIlhlnHCiFucbTMgdkZadkSVf1CbrpLjtJnqoH47gVXBwavxmXCAmXXfg1PXsJRLh17d0MCpdD15g1fQ4deoEmT0AAAASHRSTlMAEDC/gEDvv4BAz2CfQCDfEO9gj8/vnxBQr3DfgL8wIIC/j59gUF1wzyDvEK/fEM+f7yBQ34+vmHBoMIAwj3C/3yDPQFCvr/s8OcONAAAbuklEQVR42u2de3xU5ZnHDwlJSLgT7ndQFLXitWq1brvbdXdmzkEgISQkEK65EEBYQgIUiQYWQij3gEREKgJy0YA0lA8KYrgYEFshqEiF6NLd0m21dXdbd/ez/Wef53nfc3/PZSYzCY08Hz6TmZMzM5xvfs/lvOc9zytJN6U9+FQoFPob6ZZFZA+HVLvjFoywbUDIaN+/BSQ8zw1Z7RaTSMQ34L5bACOOfA+yV/d/GwEmDx4yJIg2ZPDtd0WkPkPS+FYBTGbc7DbY5wfcgazuF/lz64f3WNDDhvfy/Awk1dGy7QHid08rpzc86M9u8wp+DwqhtnIB9giGY485fMo9Djpr7YX07cHwzSF5POjk1a1YgEOCkVkvO6eOjrp0+t23Fx/aXWZM/V3SSqsVYLBJZqxsxN4L1rEVC/DeYFNNE2F/17qmdQowORgN81MXtk5+wSjZY67f8kBrHclKDkbPXL6GjSM8EI3/cUK7du0SBNsDujn+pqUL57DSsdG+Hx0H7hYw2sAIASbpFnv3zcjIyOEGTz13HyK5VIBifgGh3SnYM8m2VzuHD+rqCjAQNUXGubPLyRlJtoCMPUeO4bux28WQgJP52tHh980D8DYveAvS0tKeNhi8TEOQORnhEQy5BUBHgIHuxt0GemI2bh3UHACHu9Jj7EYxy87mT0YRxgWuDB34OQXAgIvpe3Xx3sdZvrEB2MMRH9IjdtnZ2RMm5Go2YcIEIokQ3Rj2CoefK8AkP3u1EEBneguQ3vWznyC65TNMtnw5YszWGGZ4E7yD87snEoDaAfZz2aeb6IOGxhqg+MiZ+EbJN17/CZg8YzTZGDL2fAlQBIgkxLQ0J4S9bPycJyQE/BAMf58YA+zhhA9d9xzSexns9PUxY8YbjXEEiIwhIhQ7spWfSwVtKErQuos49LSWKF0EGaM5AYryRwbDl52d++vzAG8L2fl1U422bjFRRIagQ0K4YGSOM8H7fVwJth9NdxscwSEPtG0zA+wXS4C3CeVH+CZA2FsyuhHhPffcc5WV4/LyCgsLa2rgoTAvb9IkpKgz5AgzHAje5+dKuuBo+lnTiOiIu7sDDMQQYJyD/Bg+jHnjZYRXuW/fvjMT09PTfyE3NJyFn+k1NTVAUWNICCGdCEVomE8khQnQdoh+sEh+KsToABTJbwHKT8W3buqk60BvIdmNhkVvoGUtmz59OqOIDBczhBgLnxYmkx4P+JuO5QNgws0FUOy+KD/ENx7x5Y0lem8YbWVmPtoyhFhTCAzXIcIlM0iEAjfWZsI8LDUV4J2CwtC40zAhwG4xAtjDgV/u8iWAb/FUwFdzY55G7xlun0wGywRjDEmGhFAVYY4DP68hGNHRJJk3thMecTvLmIKfc5QoAHzIHv6gch7F5bdual5hTXp6gw5vHtkz8rhxJSUlRUVEkTMEhODIXIQWgho+zyFAHwDv9ACY1JwABfy4+6L8JiG+6fkNjN48zT4ZizYOrKToBDDUEY6nSGglGMZcQB8AE4QDVQleAL8TC4Bifsx90XnTp0/Pz89s0OEtQvt4IjNGERmiDBEhEyG4sYlgOHMpRUfTz/mQuwkAWlw4yfksuckAHxLzQ/dl8kN8kyc36OzI5qBloRFEYEgIQYUgwnVWgr8Nay5qeGWMbbBU8KZYArQKUOWH7ovyW5afmTm5qOTSZo7v0KFD28EOTyObNWsOZ4gyRIQgwkJy4yXLNYK/0/n9YUhTAd7pcZYmBDjI8Sy5qQCtGThH47eO+KH8ikrGjR078STRA3anTpWXl/98yxQyFSIwZAhJhODGGsGM4B90fle8L3Z6AWSvh/ohaAiJMQMoqF84Pwp/gO9EybhxYydOzJpTD/SQHTN57dq1B2fOnEkQCSGqEAnmoxsTQebF/26YTO7ncrHkPjQqHo3p7gGwn6lAjB5AWwDE+oXxK0xPX4by4/hmTZv2CZHbvHn+fPmqXIpWXFwMFJHhrFlMhRgKNYIUB7+w8Av2Ch+g4KpRT68LJmaAktM5SlQBMn6Qf4kfhr/J5L1ZWYBvypRLBA/t8Fxuq5EiMeQqZCLUCWZfsfLzlqBh4ADNz3i9qUZpRoA9bA7MAiDLH6Q/jd/MrXsZvo1gZWhLly5lDIvXziRHJoIlnCBlkhmG+IcB0NeU6oCPwWZfBI0AVR9OiCpAuwBZAFT5lajuO2Xmwa27OLwVYAWqlSHE1RwhiNCgQaxmjPEv5HfSjAs/48XhLp7XPU11tdtgTqQALVOwmANT/azzI3xTZq4t3k/4VpBVvTCbmcpQQ0i5hBEsnLT4CyE/jzkzzgC7eu44tHkBWh0YKxgKgEZ+6L0Hi0tX/4rRqwLbVSGXn5PlrVu3cohlhBBCIRKcqBLMuyLm5yVBZwF28d5T+OvuhhPlntEDOFjswIunFhr5ofxKV8+9tJHRqwC7IJM9i4YQEeHcuSRCHgiLJucvSzfdDfc7XzNmwrguTNbV58wEySLB6AC0l9DMgfNqeP6YmDWLy29uGeHbVVGx96TM7fzzaMiQEJIINYInMvNN/L4IY+pgIAyCw9xkGmOAPawCpAxMAXBZPuMH+oPoVzp3aZm8AsV3WTZY7clNYMSQEIIfc4KYi8eZb8ccGS2APibIdBEDHGSIpFEBKBCg7sAnSjR+q+cu3QT8jsoWq62VX3jhBWLIERoJmvmlWcanbwujkE4wV4I9rbt/x8fMBJO/RwngXQIBLtEcGAPgnFmcX8H5q7Ldqi998NMXNIQWgmZ+cEaX41+CHlWzD8m6XyMZGB2ADgI0ODAEwIPA74wstLPHfkpGCFGEBoL/a+b32+xRJMHLjZECNB7iMK9pltqZsdNFplgApBRMAiQHVhMI8CvbJAa4/52P3n77bR0hujERXDvzKzO/f5yRO0p+82zDP61s+Fz9vkfCBOh1jEIJOk1niArAXrYaUBMgOTAFQMgfZfUrqk6KAK6S3//oo49UhDrB1aX/YeZ3ZcySXPn1D99cCSb7kWAkAKXugoLayaeHRgNgD7sAMQWTALkDQ/0H8W8FFC8ignVvffDBB+8DxLe5CCkQAsGvLf0Q1o259vFPOMCP37wWI4AWN3YFGIgGQFsKwRpQE6AaAJcW1GPxd1YQAi/Kn332GTIkFeoEr1j45U2d/PLLAJARXLnyEvvKh8IEaL6Qqb9IEJ8biz9oWKwAYgphAqzRIuCUxn3gwEdqqyo27BAI8OjFixe37SCEJEJOcLa1H0d6zaQbWwAgl+DKlys9JSg8mgRPgPbZRY6FTb+mA0wWpxCsATEFZ80q2Fs+//AGOP2oqNggCoE7wC5ePEkqfF8j+KyV3++m1zQ+BwBVH15ZGYwMoOQN0MghwXNktokA73XyYKgBSYCflm+ev3Ej8jssTMI/AwOEJw8hQpXg81Z+X8g3FlY+t2UL+TDg+0SWYwhQsmQRP8OwkQIU5WA9hUAKri+fP38jClCoP/nSgQMc4blqnaCtoc7kj88s3Fe5RfXhRl8DCsKjGRgewKTmBbgAczB5MBTRLIXMnIcAnfjJL7574EViCKHwwqU6+SwQfNvGb1xRZj4AVH34U+OXPhIewG4OANuFB7BnTABSCFQ9mFII1oBMgGJ+8rtoBxAhRMK6urpVez762sbv06v18pGF6MOUh6/l+DubEx5NkgPAnuEBlKIE8C5RCFw3qUbz4Clrz5AALzkA3LlzJxAEHZIb162qW/VnG78/LZo375k33lB9+MO0pgAMOAB0iJPtYgxwsBmgIQRCDiYPLi7ddHhF1VsO/Bp37iaC73INXqy7YuP3m0VmgK+njWw+gAlOHzQsOgDFVWAhVdFUBMJZ8NL6Fccc+MnHjx/fvZsYcoI2fKE/bj+kAaQg+KllTCvSyUVdzSPRjsW2918iigBZFUinITiOBWchUETLVU785HeQ4PGdnOCOHXZ+IfnUIZQgAKQg+JPKUU9HDvBOyz2tCcIsEmgpgNYcgiGwdG5BgbiAwYf3wIAhI/izAwJ+xfWntqs+TFnkUyvAZN8Ah7kechfHQUMvP48WQBxLZWU0zyEUApeK+a3fBgQbt+3ZhgR376QwKOA3t/Qgzz/14MNYSp8e9fSCHF8tywI+xvRtNwmbxxIGRQzQ9f48d4CUhCmHXC3AEHhBxG/Ptj17Lslr9uzZtg0R7kaEAn5fzS1dO2XaHLq6VM+yyNlsC8AhTbkmMsjPTmKAPWMBEMcC12lJ+Oipnx858qKIH+Dbs+alS+vXAEIQIWrwZwJ+XxcwgFkAcDNPw7nRAdjdz04JbgClaAPM4ABZFQPnIR8fOgUnwmJ+a15az20NI/iNgN+V2QVLS4tnTpsFAOt5HSMDQJ+FoL+LcmHdbNgcAHkZCFXMeQR4Uhj/9qxh9F55Zf16BPjebgG/0JlGDWA9q2Mun10ZHYBd/OzlPiph9uFYAPzF9lPlostw21QBvgIAvwSCAFDE7/dVh1WAi3ghuHjMaAQ4sqkAfe3WxQOgFGuAWbtO/VJU+70H/F4ifutf+UvoS3Div4j4/WvFrrdUgJ/Oe+ZXlwHgjagA7Ok5r8N9XnCzAZz2sujk471tHCB68B9DoS/XiOJf6EpFRdWK9znArfOeyZQB4LUoABxk21HUNqGd5A1wWIyTSNYbAn67AeAFHSCi+j8Rv9CGDQBQVmPgPjnz7MJ9MnfhHP/zO3yaOP028VMiLGN0gKLRl93vIEDuw6+sDzlZbS0CPKKXMTgiCACX+C9jwrWEnjgTuN3AqP0ZIgFoLKTr7fxeMwFcv/4bF34IcKOhkM5v2NcwPpYAW8Icz0TgVG6RDd/JFwEgxUAASAivOPD7fTUIsEJeBQo8yABOzlz2y4bT48fMyPV9KvdXCNAwmFA0zj74dwAAvnUcszARBIAO/P6tGj24cfbVsrk4RSZr4tgSAJieN3X8aARovn0zudUANAxn5WdOtl8B3nH4AADkPswIOgA8BgLcK2+FE5HVAHBW1sRxJZNxri8CnOB7OOuvEaA+oCoLAZIPv4M+TAS/ceRXKz+P07TmqmdyRQCwJm/dGASY1ooAftdpSP+cFd+mI0eP7iAfRgkSwTUviUPgfgSIsxMKyrQcUpQJACdBFbMcp7gFWw3AOFEhCFnENhfw+Xl1r+oSBILoxWuc+W2g+THkwRgCIYfkp9dMZUk4BmXgTZSGKYuMtfLb+qxct/fVo4eZBJEgDmhdEPFbRQAvowBVDzaGQFsS9gIYFx/fJhZHnqgw68BetleUKARBdWqMlV/B7Gcv7N+7l0vwrd3gxYBwmygE/hcT4GsoQObBFALRg/OmjhHkEPcysDcdZAz4KZq1xZdtlKgAHMmyiAXfubKC2Vtf248S3HGAEzz+zjvCIZj/WbW/uvpFeT6lEN2DT6AHayHQBDDO7b+XoihxyfwYo2vJUrzCQbIfqVEAyK+sW/jVr15aUDC74RgSBAkiQUS4+/hxUQW4/5j8J5yexQWoefCy9EI1BPrPISnMwSIAmIw44hR3DbLHNuy5Jz/xDg/ZS+nLZn4NND919uVqJHiUEUSEwFA4CCPT9LZnaZYvClDLweGHwE5KIv7orCRH4KHIzw1KB/qtorSHx7ZK72QvgCmKz0rwF2Z+FWyGdMHJ2mPHVhFBiIOEcOd/CpPwFXWKqiZAPA1RPdhWBQY9NSKFGZ1SFB43FSXFZbd4Jd74DYp3wPTpw2Z+/41TfEvBieUNIMFVkEiYCAHhaw6nIV9p/DQB8hws8uDH3ASYwoFEkh7g/e1FeNtruyUnKqlcjJ3iE6MCEHzYzK+UpneAE5ddrdhQCwTriCAhfFE0Cn1Yn2S+uvjgFKMA2WlIGB6sycMHwD5tzRHQ4S2QbBMNUPQv0vZv38HyFr7B8T9hvlEpx8wPBwX5bQ4XqlSCrwJCYmidBAMVzDadHzrwlFlZqgAxhYzOzY7Ig1NZ3cYZ6KiSexskoqpWj2yS7Q1tjNCUPtq7U5Q22uZE9hdob96Q6vxXNB2OeRq+dqPNweLVZ1YQQcwkDOGO35imYO2FAvrwxk0qP9WBKQUzAbIUYjkNGe4DYGfNiyRTZkgx0lDU6hiOvI/xaDsZXhjEjG9JZNj6KFZc+Ns2lg19/ACsNPFbpt7rik68iKap1laTCBHhUQ3en1+tq1u1f/+xRv02mzLD7YZFmaoAw0shJq9hUSiZM0jsQ6mTMkUcxjD6oQLiOuug6IVefGdbMNCraEw2ChdoSif2d0hVPyklziuMGOfpm/jxu635zZql8zdurNq1AVIJIIRksvdVLr1X914kfNX1+q1yyO+gyg9HAlkEtKcQvwB5WuVPFF6kKKlxPN0SPzXAKYm4MVXh3qy+wUIBX7ThbwdifehsrgMhTWEb9Q2JrnHYgR+731+73Xottpqo2kUiPLYfZFgHKWNvHc7qXYX8Di8y3+6q8ithd/1zAVrGUl09WFIMkQkPAIWmtFfiyPskeKa0YbLUpMZ0lthX6QsHn8LgpCp9+RtIwVIn05+nPf6uPbGNI1xQulPFBz/5hlTcgHu39QJo4kcdO5YZblgvJ4J4u0htNWO4H8mBHTtWXV37lvmG64NTKACOLTnBOndQ85hRYQnQoJjEREDSiQ6DWLahA+VHq1DA6sRdnO0jYVpQVPWxN6QQ4D5a3IQyUAK2neHvEt9WPTFuw8Mp1NV9jBvwSbyHBBuNo/ej1Rs2WcsEIHikfPN8cmNAyBhyq8bBg2PPq20TqGsC6I/4saYJ1DtmiSAFuwPksSyROWlnkF5nlWpnCFtt2EiKwiQJPoghTeH7MAjGNyTSgItB1ank5ZDhld70DR3UcKBqvpO+Ab/CuSpnvY9vGPidz2XDgkaCcjkSxBnnu8CRwWq5bdiwYx7rmMDlx/IH58dP4ngNaBGgxykaI5NoUqMejOK1ks6QPPiZreISSY3nK/Fscxwrp+nz4tQyGySobUjWhr2cJWj0X63rBL/pGgm+dooTBD/eRTLkVlFxjsTH+56UFhernWN4AOQlDAowjPvVSYOJWoWhDtroFZ49rndK7SsZiPCaOs5Y0XQy5uDO5g8yxjiKBm2tZ+VuAI38tL4neYa+T7N+vr1cRUiNT3YBxqqKqpNnzuitd9S2Mcx/kR93YMogVgH2aObRY9P5WDQHGXuZAmAj67xDt10XcoLoxecOYde7zQwh6x6DHXgYO6Jn6f3E+TEHphImI9iSg/ksvUQydtg2HkJgW5eAEzxtiH9q76clrPnddN56cWJW46Ht23nfO9Y9CymeR3ZljJ7efWyikR9r32Z34GCz8wsTYN94DJO9/Vxb0vl9bGofqHbPOoEENy1adIi1XuTd75CiTK3biJ6Kb47WOMvIz+bAzS5A/yNjbbFy7hDGKKTOr9LSwJIR5IHwCHb/5O0r1RaCV7UOjDM1fGrzO5Wf1gTU30ppsbH2ium02GmvRIUPXoRjOr9/sbVQnaq1sCwZV0/9Z40dQDdvvlCs9QCdpuKzNLA08LvWcgJM9ro6lcpOSiKwuzV+PxY08V3HeiBnTgaEvAOy3oO2vHw2tU9V6XF85haqehtfuSUvB3fq684u4itXuv6sXUCtbXwzf601keZdkLdvn2brgwz4LE18s9VG0nLLlTCOiYIqwt5N+IQfiviZCFIjUHDj/Mx63sdc68N96NAbgG4Og6d2keby09pIa/yMAG8Cdm3ilYi9VrfvifkZCLJASL24ZWykb+wEv2iRTOh4G3PehJtawedNpUbmy42t4OVrvq4GN4OxC/bO4wMROLCgmb6+Fga5cfp0dSUHfS2CefJYdTWCErWT/jImP0ErffmmEGBfPvLSOxofpvOzNLI0L+eA/fTzamrS+WIi+moY82RayKGELeWgruVQyORnW8zhRovzi+fwUqL0eQZ+ggVtTAuKEEJ5oWU9llNFRYiOVmTJ1/Bx+S3PFS3I4qcHd0zhKX2j9pF3G/mJCKpL2qAICWGlZUWgI5nMCN50w7JKo50WtPFo2RYj66DCi4/mp6oJ5G7xRWJtUaBRjbkzljCE5yq1NakIYD5bTongAT1aUYnhY/IzLakkt5ADqzPZwj/F8OfAGj/pduGyXmnXIBLKgBAXprqBq6JpEBemq6aui7ZYXRZtwiiSX0b4i9ZH29i4bFPLFWd+jxs2DQ+K3Rj9OHcGxsJf8XX52OJy12vI1JX5Fmuryk3IdlzbsPn9N1oZ15ufeGnSHMol6tqGp7fwxQ2R4vVJaOrakGPUdQ0nsFX5hAsbtkwCiYH9nYifmGCGjjB3hqwtrwl2fTGacXXS5bRKLuETLq0Z11r4Pc74fc/lMrENITHM4gu8EsbrY/QVcvX1cR3xya2Hn+TEz22BZra8Na0wTBRPX76O4JawJZoRHqOX5rjI9UOtjN+jjteYLFZpWOP66utoiPCscZFwvkY40hM7b2vyX15B3+10iSTossr6AvlDMGT4ebZq+ir1I91WqW81+UP6kfH8Q2DDgy4MP3/zzTeR4YfnnzZYWhotT+9ML/z6JVVxnhztMLbcHW+1DiR1MfV44ovX4D3sSYFoOrDLDrcHXWzlypXA8NwnDWfHLkBuYCNHesCLpP5T+nZWwgRILckCAwNhAjQsFRjqKOnP4GFARPys/bnNdo66GGfkGCwjwx1eRCPQSgcpnqZP4aSNthKfWyTxmW64le03qJ0BFrY2TmINjpEn/kwYin16Al3hISlwZ0DqxigO1Be+AXr3AbYnBxA2ADfgSXyGLwd8HzZ0NP/HnhAWgJqNeNw9EAaDpwHgPwfDstsjO4HAe0TaI7N4RQWYTBOa4SGV3xrT3dAjAYDRA/4EyQUIaEJgKP74DrBLwl/00ztMdNEAwr+HQw+HmALhSYhkCQA7wsMdAgHe7ejeI/iT7zrS+PHplXJ4/CIe+VQnSJLX0mTwtmxKIA5O9dFAaOscWgBKBBC7DsLLBA6QLwlmaIOArJ6UQvd15ACle5Bdx/4IEHma/1v/4ObAIwyl4SOOPCqbhZ9CUypT4uMtAOPj4+BnsjqFyNJlrJsdYNfAMAPAgaadDQpkkkOA/UNuAInfD5yyywhf9UxYFunwn0JzKpkL91FdGC+QK3GSYZqGYRUvWjzTDjAwDLw3MHQYAUzqygF21fvIAK4BBKs/d2HJCPCOkN8MYt/+WBT4NWUQhTxVwUtAKMNERVJnPBrnufTUlukLDISIyAFSEiG8CVrbfATYRdBcMYSikx4I3afGwKf0GCg9GXrKDvBRcXH9hH3rvU3EF7sJHJHeBSvo/9K/v/Rgf0l66oF7pP4PggzveLhjf1Qj/Abyhz8B3u0YF5uC7/aY4YMYGNl92GGvOxryA9BpYIFZj5sPX4vZ3wsAjnAta8iG3FTO2/KjMEaAP/Q+LaELC+HRGy61UpPN41iPspcjfL33Lr/0ekit17Qr6d/70RPqRbkRYbx/yLcy8IkAqvaD8D9jsBO7e5OlVm8jzPh+1IQyYrCuxiGD75K+LaZPZ3MY0L9lXkbXM3/wt7dARGL/D7ZxXd3U+lPKAAAAAElFTkSuQmCC"""



def run_report(input_file, output_xlsx, output_html):
    """Process a user-uploaded BSNL FTTH Excel workbook."""
    SRC = Path(input_file)
    OUT = Path(output_xlsx)
    HTML_OUT = Path(output_html)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    HTML_OUT.parent.mkdir(parents=True, exist_ok=True)
    if not SRC.exists():
        raise FileNotFoundError(f"Uploaded Excel file is not accessible: {SRC}")
    return _run_report_body(SRC, OUT, HTML_OUT)

def _run_report_body(SRC, OUT, HTML_OUT):
    # ---------------------------------------------------------------------------
    # 1) OLT IP -> BBC Name master map (same table used in the VBA module)
    # ---------------------------------------------------------------------------
    MAP_RAW = """10.214.132.5=SURENDER G|10.214.132.6=SPANDANA T|10.215.132.172=MAHESHWARAPU RANJITH KUMAR-WDP|10.215.132.141=KIRAN KUMAR ENUGULA|10.215.132.60=MAHESHWARAPU RANJITH KUMAR-WDP|10.215.132.177=A PRAVEEN-JAN|10.215.132.163=A PRAVEEN-JAN|10.215.132.30=MAHESHWARAPU RANJITH KUMAR-WDP|10.215.132.33=VINEETH GUNTI|10.215.132.133=MAHESHWARAPU RANJITH KUMAR-WDP|10.210.49.19=A PRAVEEN-JAN|10.215.132.145=SURENDER G|10.215.132.101=KIRAN KUMAR ENUGULA|10.210.49.24=VINEETH GUNTI|10.210.49.25=SURENDER G|10.210.49.27=KIRAN KUMAR ENUGULA|10.210.49.35=MAHESHWARAPU RANJITH KUMAR-WDP|10.215.132.217=VINEETH GUNTI|10.210.49.150=SURENDER G|10.210.49.194=SURENDER G|10.210.49.69=CH MADHUKAR RAO|10.215.132.43=SURENDER G|10.215.132.42=SURENDER G|10.215.132.198=SURENDER G|10.215.132.212=MAHESHWARAPU RANJITH KUMAR|10.215.132.55=VIKRAM NALLA|10.215.132.208=CH MADHUKAR RAO|10.215.132.210=VINEETH GUNTI|10.215.132.92=VIKRAM NALLA|10.215.132.23=VINEETH GUNTI|10.215.132.85=VIKRAM NALLA|10.215.132.118=SPANDANA T|10.210.49.34=SPANDANA T|10.215.132.46=MAHESHWARAPU RANJITH KUMAR|10.215.132.169=SPANDANA T|10.215.132.106=SPANDANA T|10.215.132.211=SPANDANA T|10.215.132.44=KIRAN KUMAR ENUGULA|10.215.132.223=SURENDER G|10.215.132.97=SPANDANA T|10.215.132.209=SPANDANA T|10.210.49.29=MAHESHWARAPU RANJITH KUMAR|10.210.49.10=SURENDER G|10.215.132.170=VIKRAM NALLA|10.215.132.38=VIKRAM NALLA|10.215.132.150=SPANDANA T|10.215.132.34=MAHESHWARAPU RANJITH KUMAR-WDP|10.215.132.157=CH MADHUKAR RAO|10.210.49.204=MAHESHWARAPU RANJITH KUMAR|10.210.49.8=CH MADHUKAR RAO|10.215.132.216=VINEETH GUNTI|10.210.49.15=VINEETH GUNTI|10.215.132.75=VINEETH GUNTI|10.215.132.87=VINEETH GUNTI|10.215.132.94=VINEETH GUNTI|10.210.49.190=BHIKSHAPATHI MANCHALA|10.215.132.132=MAHESHWARAPU RANJITH KUMAR|10.215.132.226=BHIKSHAPATHI MANCHALA|10.215.132.144=BHIKSHAPATHI MANCHALA|10.215.132.146=BHIKSHAPATHI MANCHALA|10.215.132.143=CH MADHUKAR RAO|10.215.132.129=BHIKSHAPATHI MANCHALA|10.215.132.130=BHIKSHAPATHI MANCHALA|10.215.132.201=CH MADHUKAR RAO|10.215.132.142=CH MADHUKAR RAO|10.215.132.171=BHIKSHAPATHI MANCHALA|10.215.132.48=CH MADHUKAR RAO|10.215.132.151=BHIKSHAPATHI MANCHALA|10.210.49.20=BHIKSHAPATHI MANCHALA|10.215.132.113=CH MADHUKAR RAO|10.215.132.89=BHIKSHAPATHI MANCHALA|10.215.132.127=CH MADHUKAR RAO|10.215.132.206=BHIKSHAPATHI MANCHALA|10.215.132.140=BHIKSHAPATHI MANCHALA|10.215.132.84=MAHESHWARAPU RANJITH KUMAR|10.215.132.123=CH MADHUKAR RAO|10.215.132.120=MAHESHWARAPU RANJITH KUMAR|10.210.49.5=VIKRAM NALLA|10.215.132.25=CH MADHUKAR RAO|10.215.132.47=BHIKSHAPATHI MANCHALA|10.215.132.117=VIKRAM NALLA|10.215.132.228=SPANDANA T|10.215.132.58=SPANDANA T|10.215.132.69=SPANDANA T|10.215.132.215=MAHESHWARAPU RANJITH KUMAR|10.215.132.83=SPANDANA T|10.215.132.77=SPANDANA T|10.215.132.74=VINEETH GUNTI|10.215.132.219=MAHESHWARAPU RANJITH KUMAR|10.215.132.173=A PRAVEEN-JAN|10.215.132.32=VINEETH GUNTI|10.215.132.149=SURENDER G|10.215.132.218=SURENDER G|10.215.132.119=VINEETH GUNTI|10.215.132.10=MAHESHWARAPU RANJITH KUMAR-WDP|10.215.132.104=MAHESHWARAPU RANJITH KUMAR-WDP|10.215.132.147=SURENDER G|10.215.132.168=KIRAN KUMAR ENUGULA|10.215.132.15=BHIKSHAPATHI MANCHALA|10.215.132.131=CH MADHUKAR RAO|10.215.132.134=SURENDER G|10.215.132.165=CH MADHUKAR RAO|10.210.49.65=MAHESHWARAPU RANJITH KUMAR-WDP|10.215.132.128=A PRAVEEN-CHR|10.210.49.200=MAHESHWARAPU RANJITH KUMAR-WDP|10.215.132.16=SPANDANA T|10.210.49.59=CH MADHUKAR RAO|10.215.132.237=SURENDER G|10.215.132.73=SURENDER G|10.215.132.160=SURENDER G|10.210.49.191=KIRAN KUMAR ENUGULA|10.210.49.30=A PRAVEEN-CHR|10.210.49.70=MAHESHWARAPU RANJITH KUMAR-WDP|10.210.49.26=KIRAN KUMAR ENUGULA|10.215.132.174=BHIKSHAPATHI MANCHALA|10.215.132.176=MAHESHWARAPU RANJITH KUMAR|10.215.132.24=VIKRAM NALLA|10.215.132.240=VIKRAM NALLA|10.215.132.189=CH MADHUKAR RAO|10.215.132.254=A PRAVEEN-JAN|10.215.132.225=VINEETH GUNTI|10.210.49.23=VINEETH GUNTI|10.215.132.224=SURENDER G|10.215.132.64=A PRAVEEN-CHR|10.215.132.253=SPANDANA T|10.215.132.155=SPANDANA T|10.215.132.80=A PRAVEEN-CHR|10.215.132.114=A PRAVEEN-JAN|10.215.132.40=A PRAVEEN-JAN|10.215.132.99=A PRAVEEN-JAN|10.215.132.31=A PRAVEEN-JAN|10.215.132.88=A PRAVEEN-CHR|10.215.132.79=A PRAVEEN-CHR|10.215.132.78=A PRAVEEN-JAN|10.215.132.90=A PRAVEEN-CHR|10.215.132.244=MAHESHWARAPU RANJITH KUMAR-WDP|10.215.132.236=A PRAVEEN-CHR|10.215.132.45=MAHESHWARAPU RANJITH KUMAR-WDP|10.215.132.252=A PRAVEEN-JAN|10.215.132.20=VINEETH GUNTI|10.215.132.251=A PRAVEEN-JAN|10.215.132.29=A PRAVEEN-JAN|10.215.132.164=A PRAVEEN-JAN|10.210.49.37=A PRAVEEN-JAN|10.215.132.233=SURENDER G|10.215.132.232=MAHESHWARAPU RANJITH KUMAR-WDP|10.215.132.21=SURENDER G|10.210.49.7=A PRAVEEN-JAN|10.210.49.4=SURENDER G|10.215.132.178=SPANDANA T|10.215.132.57=A PRAVEEN-JAN|10.215.132.200=MAHESHWARAPU RANJITH KUMAR-WDP|10.215.132.22=VINEETH GUNTI|10.215.132.63=KIRAN KUMAR ENUGULA|10.215.132.95=CH MADHUKAR RAO|10.215.132.214=BHIKSHAPATHI MANCHALA|10.215.132.96=BHIKSHAPATHI MANCHALA|10.215.132.76=MAHESHWARAPU RANJITH KUMAR-WDP|10.215.132.185=BHIKSHAPATHI MANCHALA|10.215.132.65=BHIKSHAPATHI MANCHALA|10.215.132.71=BHIKSHAPATHI MANCHALA|10.215.132.67=BHIKSHAPATHI MANCHALA|10.215.132.51=BHIKSHAPATHI MANCHALA|10.215.132.115=BHIKSHAPATHI MANCHALA|10.215.132.50=BHIKSHAPATHI MANCHALA|10.215.132.81=BHIKSHAPATHI MANCHALA|10.215.132.239=BHIKSHAPATHI MANCHALA|10.215.132.124=BHIKSHAPATHI MANCHALA|10.215.132.121=CH MADHUKAR RAO|10.215.132.111=CH MADHUKAR RAO|10.215.132.227=KIRAN KUMAR ENUGULA|10.215.132.122=CH MADHUKAR RAO|10.215.132.230=MAHESHWARAPU RANJITH KUMAR-WDP|10.215.132.126=BHIKSHAPATHI MANCHALA|10.215.132.39=CH MADHUKAR RAO|10.215.132.109=MAHESHWARAPU RANJITH KUMAR|10.215.132.70=MAHESHWARAPU RANJITH KUMAR|10.215.132.72=MAHESHWARAPU RANJITH KUMAR|10.215.132.154=CH MADHUKAR RAO|10.215.132.82=CH MADHUKAR RAO|10.215.132.54=KIRAN KUMAR ENUGULA|10.215.132.204=BHIKSHAPATHI MANCHALA|10.215.132.53=CH MADHUKAR RAO|10.215.132.52=CH MADHUKAR RAO|10.215.132.125=BHIKSHAPATHI MANCHALA|10.215.132.68=SURENDER G|10.215.132.238=CH MADHUKAR RAO|10.215.132.213=CH MADHUKAR RAO|10.215.132.62=CH MADHUKAR RAO|10.215.132.91=KIRAN KUMAR ENUGULA|10.215.132.107=MAHESHWARAPU RANJITH KUMAR-WDP|10.215.132.241=VIKRAM NALLA|10.215.132.19=VIKRAM NALLA|10.215.132.112=VINEETH GUNTI|10.215.132.229=SURENDER G|10.215.132.11=SURENDER G|10.210.49.195=BHIKSHAPATHI MANCHALA|10.215.132.248=MAHESHWARAPU RANJITH KUMAR|10.215.132.245=KIRAN KUMAR ENUGULA|10.215.132.246=KIRAN KUMAR ENUGULA|10.215.132.162=KIRAN KUMAR ENUGULA|10.215.132.249=KIRAN KUMAR ENUGULA|10.215.132.250=KIRAN KUMAR ENUGULA|10.215.132.159=KIRAN KUMAR ENUGULA|10.215.132.222=MAHESHWARAPU RANJITH KUMAR|10.215.132.221=MAHESHWARAPU RANJITH KUMAR|10.210.49.6=A PRAVEEN-JAN|10.215.132.105=SPANDANA T|10.215.132.220=SPANDANA T|10.215.132.18=KIRAN KUMAR ENUGULA|10.215.132.158=KIRAN KUMAR ENUGULA|10.210.49.36=VINEETH GUNTI|10.210.49.51=CH MADHUKAR RAO|10.215.132.175=CH MADHUKAR RAO|10.210.49.55=MAHESHWARAPU RANJITH KUMAR|10.210.49.40=BHIKSHAPATHI MANCHALA|10.210.49.44=CH MADHUKAR RAO|10.210.49.49=MAHESHWARAPU RANJITH KUMAR|10.210.49.43=BHIKSHAPATHI MANCHALA|10.210.49.50=BHIKSHAPATHI MANCHALA|10.210.49.45=BHIKSHAPATHI MANCHALA|10.210.49.67=CH MADHUKAR RAO|10.210.49.41=BHIKSHAPATHI MANCHALA|10.210.49.52=BHIKSHAPATHI MANCHALA|10.210.49.56=BHIKSHAPATHI MANCHALA|10.210.49.47=MAHESHWARAPU RANJITH KUMAR|10.210.49.42=BHIKSHAPATHI MANCHALA|10.215.132.179=CH MADHUKAR RAO|10.210.49.57=BHIKSHAPATHI MANCHALA|10.210.49.58=MAHESHWARAPU RANJITH KUMAR|10.210.49.48=CH MADHUKAR RAO|10.210.49.151=SPANDANA T|10.210.49.33=VIKRAM NALLA|10.254.48.3=BHIKSHAPATHI MANCHALA|10.210.49.71=A PRAVEEN-JAN|10.210.49.63=VIKRAM NALLA|10.210.49.64=A PRAVEEN-CHR|10.215.132.59=A PRAVEEN-JAN|10.210.49.72=VINEETH GUNTI|10.210.49.75=SURENDER G|10.210.49.180=A PRAVEEN-JAN"""
    MAP_RAW += ("|10.210.49.199=BHIKSHAPATHI MANCHALA|10.210.49.81=BHIKSHAPATHI MANCHALA|10.215.132.235=BHIKSHAPATHI MANCHALA|"
                "10.210.49.85=MAHESHWARAPU RANJITH KUMAR-KZP|10.210.49.198=MAHESHWARAPU RANJITH KUMAR-KZP|"
                "10.210.49.157=A PRAVEEN-JAN|10.210.49.156=A PRAVEEN-JAN|10.210.49.73=A PRAVEEN-JAN|10.215.132.181=SPANDANA T")
    OLT_MAP = {}
    for pair in MAP_RAW.split('|'):
        k, v = pair.split('=')
        OLT_MAP[k.strip()] = v.strip()

    # canonical BBC master info: Manager, ManagerTarget, Area, BBCTarget
    BBC_INFO = {
        "BHIKSHAPATHI MANCHALA":               ("M RAMMOHAN", 225, "HNK", 87),
        "MAHESHWARAPU RANJITH KUMAR-KZP":      ("M RAMMOHAN", 225, "KZP", 40),
        "CH MADHUKAR RAO":                     ("M RAMMOHAN", 225, "WGL", 64),
        "MAHESHWARAPU RANJITH KUMAR-WDP":      ("M RAMMOHAN", 225, "WDP", 34),
        "SPANDANA T":                          ("B SESHU SRINIVAS", 125, "NRM", 38),
        "SURENDER G":                          ("B SESHU SRINIVAS", 125, "PKA & BHPL", 47),
        "KIRAN KUMAR ENUGULA":                 ("B SESHU SRINIVAS", 125, "MLG", 40),
        "VINEETH GUNTI":                       ("G DAYAKAR", 120, "MBB & DRNKL", 34),
        "VIKRAM NALLA":                        ("G DAYAKAR", 120, "THR", 25),
        "A PRAVEEN-JAN":                       ("G DAYAKAR", 120, "JAN", 44),
        "A PRAVEEN-CHR":                       ("G DAYAKAR", 120, "CHR", 17),
    }
    DISPLAY_NAME = {"SURENDER G": "SURENDER GUGULOTH"}
    BBC_ORDER = ["BHIKSHAPATHI MANCHALA", "MAHESHWARAPU RANJITH KUMAR-KZP", "CH MADHUKAR RAO",
                 "MAHESHWARAPU RANJITH KUMAR-WDP", "SPANDANA T", "SURENDER G", "KIRAN KUMAR ENUGULA",
                 "VINEETH GUNTI", "VIKRAM NALLA", "A PRAVEEN-JAN", "A PRAVEEN-CHR"]

    MONTHS = {"JAN":1,"FEB":2,"MAR":3,"APR":4,"MAY":5,"JUN":6,"JUL":7,"AUG":8,"SEP":9,"OCT":10,"NOV":11,"DEC":12}

    def parse_date(v):
        if isinstance(v, datetime.datetime):
            return v.date()
        if isinstance(v, datetime.date):
            return v
        s = str(v).strip()
        date_part = s.split(" ")[0]
        d, mmm, y = date_part.split("-")
        return datetime.date(int(y), MONTHS[mmm.upper()[:3]], int(d))

    # ---------------------------------------------------------------------------
    # 2) Read source workbook (header is row 3 in this export)
    # ---------------------------------------------------------------------------
    wb_src = load_workbook(SRC, read_only=True, data_only=True)
    ws_src = wb_src['Sheet0']
    raw_rows = []
    hdr = None
    for i, r in enumerate(ws_src.iter_rows(values_only=True)):
        if i == 2:
            hdr = r
            continue
        if i < 3:
            continue
        if r[0] is None or r[0] == '':
            continue
        raw_rows.append(r)
    idx = {name: j for j, name in enumerate(hdr) if name}

    REQUIRED = ["BBC Name", "CLSR", "Ont Acquisition Type", "Disconnection reason",
                "Completion_Date", "Maintenance Franchisee", "OLT IP", "Order Id"]
    missing_cols = [c for c in REQUIRED if c not in idx]
    if missing_cols:
        raise SystemExit(f"Missing required columns: {missing_cols}")

    TODAY = (date.today() - timedelta(days=1)).strftime("%d-%m-%Y")   # source export's "Last Update Time" date

    # ---------------------------------------------------------------------------
    # 3) Classify + rebuild BBC Name + aggregate (mirrors the VBA logic exactly)
    # ---------------------------------------------------------------------------
    prepared = []   # full rows + [DATE, Connection Type, BBC Name Rebuilt Source]
    agg = defaultdict(lambda: [0, 0, 0, 0, 0])   # key -> [NPC, RECON, CLSVO, CLSNP, NPC_today]
    unmapped_olt_ips = set()
    unmapped_bbc_names = defaultdict(int)

    def normalize_bbc_name(bbc: str) -> str:
        s = " ".join(bbc.split())  # collapse repeated/leading/trailing whitespace
        if s.upper() == "MAHESHWARAPU RANJITH KUMAR":
            return "MAHESHWARAPU RANJITH KUMAR-KZP"
        if s.upper() == "A PRAVEEN":
            return "A PRAVEEN-JAN"
        return s

    for r in raw_rows:
        clsr = str(r[idx["CLSR"]]).strip().upper()
        order_id = str(r[idx["Order Id"]]).strip()
        olt_ip = str(r[idx["OLT IP"]]).strip()
        mf = str(r[idx["Maintenance Franchisee"]]).strip()
        comp_raw = r[idx["Completion_Date"]]

        if olt_ip in OLT_MAP:
            bbc = OLT_MAP[olt_ip]
        else:
            bbc = str(r[idx["BBC Name"]]).strip()
            unmapped_olt_ips.add(olt_ip)
        bbc = normalize_bbc_name(bbc)

        id5 = order_id[:5].upper()
        if clsr == "ACTIVE" and id5 == "BFBNC":
            conn_type = "NPC"
        elif clsr == "ACTIVE":
            conn_type = "RECONNECTION"
        elif clsr == "CLSD" and id5 == "BFBDI":
            conn_type = "DUE TO NON PAYMENT (CLSNP)"
        elif clsr == "CLSV" and id5 == "BFBDI":
            conn_type = "VOLUNTAORY DISCONNECTION (CLSVO)"
        else:
            conn_type = "OTHER"

        d_only = parse_date(comp_raw)
        is_today = (d_only == TODAY)

        if bbc not in BBC_INFO:
            unmapped_bbc_names[bbc] += 1

        prepared.append(list(r) + [d_only.strftime("%d-%b-%Y").upper(), conn_type, bbc])

        def bump(key):
            a = agg[key]
            if conn_type == "NPC":
                a[0] += 1
                if is_today: a[4] += 1
            elif conn_type == "RECONNECTION":
                a[1] += 1
                if is_today: a[4] += 1
            elif conn_type == "VOLUNTAORY DISCONNECTION (CLSVO)":
                a[2] += 1
            elif conn_type == "DUE TO NON PAYMENT (CLSNP)":
                a[3] += 1

        bump(("BBC", bbc))
        bump(("OLT", olt_ip))
        bump(("MF", mf))
        bump(("DATE", d_only.strftime("%d-%b-%Y")))
        mgr = BBC_INFO[bbc][0] if bbc in BBC_INFO else "UNMAPPED"
        bump(("MGR", mgr))
        area = BBC_INFO[bbc][2] if bbc in BBC_INFO else "UNMAPPED"
        bump(("AREA", area))

    print("Rows processed:", len(prepared))
    print("Unmapped OLT IPs (fell back to source BBC Name):", sorted(unmapped_olt_ips))
    print("BBC names with no master-table entry (Target/Area unknown):", dict(unmapped_bbc_names))
    print()
    type_counts = defaultdict(int)
    for r in prepared:
        type_counts[r[-2]] += 1
    print("Connection Type breakdown:", dict(type_counts))

    # ---------------------------------------------------------------------------
    # 4) Build workbook: Data sheet (prepared) + FTTHDashboard sheet
    # ---------------------------------------------------------------------------
    NAVY = "37336F"
    BLUE = "1F4EBA"
    LIGHTGREY = "E6E6E6"
    WHITE = "FFFFFF"
    GREEN_FILL = "D6F5D6"
    RED_FILL = "F5D6D6"
    RED_FONT = "960000"
    GREEN_FONT = "006E00"

    wbo = Workbook()
    ws_data = wbo.active
    ws_data.title = "Data"

    data_headers = list(hdr) + ["DATE", "Connection Type", "BBC Name (rebuilt)"]
    # overwrite BBC Name column (index) with rebuilt value, keep everything else as-is
    bbc_col_idx = idx["BBC Name"]
    for j, h in enumerate(data_headers, start=1):
        c = ws_data.cell(row=1, column=j, value=h)
        c.font = Font(bold=True, color=WHITE, name="Arial")
        c.fill = PatternFill("solid", fgColor=NAVY)
    ws_data.freeze_panes = "A2"

    for i, r in enumerate(prepared, start=2):
        row_vals = list(r[:-3])
        row_vals[bbc_col_idx] = r[-1]          # rebuilt BBC Name overwrites original column
        row_vals += [r[-3], r[-2]]             # DATE, Connection Type
        for j, v in enumerate(row_vals, start=1):
            ws_data.cell(row=i, column=j, value=v)

    for col in ws_data.columns:
        length = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws_data.column_dimensions[col[0].column_letter].width = min(max(length + 2, 10), 34)

    last_row = len(prepared) + 1
    last_col_letter = get_column_letter(len(data_headers) - 1)  # exclude the extra "BBC Name (rebuilt)" label col we didn't write
    data_range = f"A1:{get_column_letter(len(hdr)+2)}{last_row}"

    # ---------------------------------------------------------------------------
    # 5) FTTHDashboard sheet
    # ---------------------------------------------------------------------------
    ws = wbo.create_sheet("FTTHDashboard", 0)
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:N2")
    ws["A1"] = "FTTH WARANGAL DASHBOARD"
    ws["A1"].font = Font(size=22, bold=True, color=WHITE, name="Arial")
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22

    ws.merge_cells("A3:N3")
    ws["A3"] = f"DAILY PROVISIONS DASHBOARD  |  Report Date {datetime.date.today():%d-%b-%Y}"
    ws["A3"].font = Font(italic=True, size=9, name="Arial")

    HDR1, HDR2, HDR3, FIRST = 5, 6, 7, 8
    ws.merge_cells(f"A{HDR1}:N{HDR1}")
    ws[f"A{HDR1}"] = f"BBM WISE PROVISIONING REPORT OF WGL OA AS ON {TODAY:%d-%b-%Y}".upper()
    ws[f"A{HDR1}"].font = Font(bold=True, size=13, color=WHITE, name="Arial")
    ws[f"A{HDR1}"].fill = PatternFill("solid", fgColor=BLUE)
    ws[f"A{HDR1}"].alignment = Alignment(horizontal="center")

    headers = ["S.No", "AGM/ Manager(MT)", "BBM NAME", "AREA", "Exclusive/Non Exclusive",
               "No. Of OLTEs Mapped", "Monthly Target", f"Daily Provision ({TODAY:%d-%b-%Y})", "Cumulative Achievement",
               "% of Achievement"]
    for j, h in enumerate(headers, start=1):
        ws.merge_cells(start_row=HDR2, start_column=j, end_row=HDR3, end_column=j)
        ws.cell(row=HDR2, column=j, value=h)
    ws.merge_cells(start_row=HDR2, start_column=11, end_row=HDR2, end_column=13)
    ws.cell(row=HDR2, column=11, value="Disconnections")
    ws.cell(row=HDR3, column=11, value="CLSVO")
    ws.cell(row=HDR3, column=12, value="CLSNP")
    ws.cell(row=HDR3, column=13, value="Total")
    ws.merge_cells(start_row=HDR2, start_column=14, end_row=HDR3, end_column=14)
    ws.cell(row=HDR2, column=14, value="NET")

    for row in ws.iter_rows(min_row=HDR2, max_row=HDR3, min_col=1, max_col=14):
        for c in row:
            c.font = Font(bold=True, name="Arial", size=10)
            c.fill = PatternFill("solid", fgColor=LIGHTGREY)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = Border(*(Side(style="thin", color="A0A0A0"),) * 4)
    ws.row_dimensions[HDR2].height = 15
    ws.row_dimensions[HDR3].height = 28

    thin = Side(style="thin", color="A0A0A0")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    def agg_for(dim, key):
        return agg.get((dim, key), [0, 0, 0, 0, 0])

    r = FIRST + 1  # leave row FIRST for the "Total OA" row, inserted after detail rows are computed
    prev_mgr = None
    detail_rows = []
    tot = dict(olt=0, target=0, today=0, cum=0, clsvo=0, clsnp=0)

    for i, bbc_key in enumerate(BBC_ORDER, start=1):
        mgr, mgr_target, area, bbc_target = BBC_INFO[bbc_key]
        olt_count = sum(1 for v in OLT_MAP.values() if v == bbc_key)
        npc, recon, clsvo, clsnp, npc_today = agg_for("BBC", bbc_key)
        cum = npc + recon
        disc = clsvo + clsnp
        net = cum - disc
        pct = (cum / bbc_target) if bbc_target else 0
        display = DISPLAY_NAME.get(bbc_key, bbc_key)

        detail_rows.append(dict(sno=i, mgr=mgr, name=display, area=area, excl="EXCLUSIVE" if i == 1 else "",
                                 olt=olt_count, target=bbc_target, today=npc_today, cum=cum, pct=pct,
                                 clsvo=clsvo, clsnp=clsnp, disc=disc, net=net))
        for k, v in zip(tot.keys(), [olt_count, bbc_target, npc_today, cum, clsvo, clsnp]):
            tot[k] += v

    # data-quality bucket: rows whose rebuilt/raw BBC Name has no master-table entry
    unm_npc, unm_recon, unm_clsvo, unm_clsnp = 0, 0, 0, 0
    for name, count in unmapped_bbc_names.items():
        npc, recon, clsvo, clsnp, _ = agg_for("BBC", name)
        unm_npc += npc; unm_recon += recon; unm_clsvo += clsvo; unm_clsnp += clsnp
    unm_cum = unm_npc + unm_recon
    unm_disc = unm_clsvo + unm_clsnp
    unm_net = unm_cum - unm_disc
    has_unmapped = any([unm_npc, unm_recon, unm_clsvo, unm_clsnp])

    n_detail = len(detail_rows) + (1 if has_unmapped else 0)

    # ---- Total OA row ----
    ws.cell(row=FIRST, column=1, value="Total OA")
    ws.cell(row=FIRST, column=2, value="WGL")
    ws.cell(row=FIRST, column=6, value=tot["olt"])
    ws.cell(row=FIRST, column=7, value=tot["target"])
    ws.cell(row=FIRST, column=8, value=tot["today"])
    ws.cell(row=FIRST, column=9, value=tot["cum"])
    ws.cell(row=FIRST, column=10, value=(tot["cum"] / tot["target"]) if tot["target"] else 0)
    ws.cell(row=FIRST, column=10).number_format = "0.00%"
    ws.cell(row=FIRST, column=11, value=tot["clsvo"])
    ws.cell(row=FIRST, column=12, value=tot["clsnp"])
    ws.cell(row=FIRST, column=13, value=tot["clsvo"] + tot["clsnp"])
    ws.cell(row=FIRST, column=14, value=tot["cum"] - (tot["clsvo"] + tot["clsnp"]))
    for c in ws[FIRST][:14]:
        c.font = Font(bold=True, color=WHITE, name="Arial")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.border = border_all
        c.alignment = Alignment(horizontal="center")

    r = FIRST + 1
    prev_mgr = None
    for d in detail_rows:
        ws.cell(row=r, column=1, value=d["sno"])
        if d["mgr"] != prev_mgr:
            c = ws.cell(row=r, column=2, value=d["mgr"])
            c.font = Font(bold=True, name="Arial")
            prev_mgr = d["mgr"]
        ws.cell(row=r, column=3, value=d["name"])
        ws.cell(row=r, column=4, value=d["area"])
        ws.cell(row=r, column=5, value=d["excl"])
        ws.cell(row=r, column=6, value=d["olt"])
        ws.cell(row=r, column=7, value=d["target"])
        ws.cell(row=r, column=8, value=d["today"])
        ws.cell(row=r, column=9, value=d["cum"])
        pc = ws.cell(row=r, column=10, value=d["pct"]); pc.number_format = "0.00%"
        ws.cell(row=r, column=11, value=d["clsvo"])
        ws.cell(row=r, column=12, value=d["clsnp"])
        ws.cell(row=r, column=13, value=d["disc"])
        netc = ws.cell(row=r, column=14, value=d["net"])
        if d["net"] < 0:
            netc.fill = PatternFill("solid", fgColor=RED_FILL); netc.font = Font(color=RED_FONT, bold=True, name="Arial")
        else:
            netc.fill = PatternFill("solid", fgColor=GREEN_FILL); netc.font = Font(color=GREEN_FONT, bold=True, name="Arial")
        for c in ws[r][:14]:
            c.border = border_all
            c.alignment = Alignment(horizontal="center")
            if not c.font or c.font.name is None:
                c.font = Font(name="Arial")
        r += 1

    if has_unmapped:
        ws.cell(row=r, column=1, value=len(detail_rows) + 1)
        ws.cell(row=r, column=2, value="(unassigned)")
        ws.cell(row=r, column=3, value="UNMAPPED / OTHER *")
        ws.cell(row=r, column=4, value="-")
        ws.cell(row=r, column=6, value=0)
        ws.cell(row=r, column=7, value=0)
        ws.cell(row=r, column=8, value=unm_npc)
        ws.cell(row=r, column=9, value=unm_cum)
        ws.cell(row=r, column=10, value=0); ws.cell(row=r, column=10).number_format = "0.00%"
        ws.cell(row=r, column=11, value=unm_clsvo)
        ws.cell(row=r, column=12, value=unm_clsnp)
        ws.cell(row=r, column=13, value=unm_disc)
        netc = ws.cell(row=r, column=14, value=unm_net)
        for c in ws[r][:14]:
            c.border = border_all
            c.alignment = Alignment(horizontal="center")
            c.font = Font(italic=True, name="Arial")
            c.fill = PatternFill("solid", fgColor="FFF3CD")
        r += 1

    last_data_row = r - 1
    ws.cell(row=r, column=1,
            value="* BBC Name could not be matched to the FY2025 master table (new/renamed OLT or BBC not yet on the master list). "
                  "See notes below.").font = Font(italic=True, size=8, color="808080", name="Arial")

    for col_letter, width in zip(list("ABCDEFGHIJKLMN"), [6, 18, 30, 12, 14, 10, 10, 10, 12, 11, 8, 8, 8, 8]):
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = f"A{HDR3+1}"

    # ---------------------------------------------------------------------------
    # 6) Charts
    # ---------------------------------------------------------------------------
    bar = BarChart()
    bar.type = "col"
    bar.title = "BBC Wise: Monthly Target vs NET"
    bar.y_axis.title = "Count"
    cats = Reference(ws, min_col=3, min_row=FIRST + 1, max_row=last_data_row)
    data_target = Reference(ws, min_col=7, min_row=HDR3, max_row=last_data_row)
    data_net = Reference(ws, min_col=14, min_row=HDR3, max_row=last_data_row)
    bar.add_data(data_target, titles_from_data=True)
    bar.add_data(data_net, titles_from_data=True)
    bar.set_categories(cats)
    bar.width, bar.height = 24, 12
    ws.add_chart(bar, "P5")

    line = LineChart()
    line.title = "BBC Wise: % of Achievement"
    data_pct = Reference(ws, min_col=10, min_row=HDR3, max_row=last_data_row)
    line.add_data(data_pct, titles_from_data=True)
    line.set_categories(cats)
    line.width, line.height = 24, 12
    ws.add_chart(line, "P28")

    wbo.save(OUT)
    print("Saved workbook (before master sheets):", OUT)

    # ---------------------------------------------------------------------------
    # 7) Editable master sheets: OLT_BBC_Map + BBC_Master (mirrors the VBA's
    #    EnsureOltMapSheet / EnsureBbcMasterSheet, so this preview matches what
    #    the macro creates on first run in Excel).
    # ---------------------------------------------------------------------------
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.worksheet.datavalidation import DataValidation

    ws_olt = wbo.create_sheet("OLT_BBC_Map")
    ws_olt["A1"] = "OLT IP -> BBC Name  (editable master list - add new rows any time, then re-run FTTHDashboard)"
    ws_olt["A1"].font = Font(bold=True, color=WHITE, name="Arial")
    ws_olt["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws_olt.merge_cells("A1:B1")
    ws_olt["A2"] = "OLT IP"; ws_olt["B2"] = "BBC Name"
    ws_olt["A2"].font = Font(bold=True, name="Arial"); ws_olt["B2"].font = Font(bold=True, name="Arial")
    r = 3
    for ip, bbc in OLT_MAP.items():
        ws_olt.cell(row=r, column=1, value=ip)
        ws_olt.cell(row=r, column=2, value=bbc)
        r += 1
    olt_last = r - 1
    tbl = Table(displayName="tbl_OltMap", ref=f"A2:B{olt_last}")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws_olt.add_table(tbl)
    ws_olt.column_dimensions["A"].width = 22
    ws_olt.column_dimensions["B"].width = 32

    ws_bbc = wbo.create_sheet("BBC_Master")
    ws_bbc.merge_cells("A1:F1")
    ws_bbc["A1"] = "BBC / Manager / Area / Target - editable master list (add a row for a new BBC, then re-run FTTHDashboard)"
    ws_bbc["A1"].font = Font(bold=True, color=WHITE, name="Arial")
    ws_bbc["A1"].fill = PatternFill("solid", fgColor=NAVY)
    bbc_headers = ["BBC Name", "AGM/Manager(MT)", "Manager Target", "AREA", "BBC Target", "Display Name (optional)"]
    for j, h in enumerate(bbc_headers, start=1):
        c = ws_bbc.cell(row=2, column=j, value=h)
        c.font = Font(bold=True, name="Arial")
    r = 3
    for bbc_key in BBC_ORDER:
        mgr, mgr_target, area, bbc_target = BBC_INFO[bbc_key]
        ws_bbc.cell(row=r, column=1, value=bbc_key)
        ws_bbc.cell(row=r, column=2, value=mgr)
        ws_bbc.cell(row=r, column=3, value=mgr_target)
        ws_bbc.cell(row=r, column=4, value=area)
        ws_bbc.cell(row=r, column=5, value=bbc_target)
        ws_bbc.cell(row=r, column=6, value=DISPLAY_NAME.get(bbc_key, ""))
        r += 1
    bbc_last = r - 1
    tbl2 = Table(displayName="tbl_BbcMaster", ref=f"A2:F{bbc_last}")
    tbl2.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws_bbc.add_table(tbl2)
    ws_bbc.column_dimensions["A"].width = 32
    ws_bbc.column_dimensions["B"].width = 20
    for col in "CDEF":
        ws_bbc.column_dimensions[col].width = 16

    # dropdown on OLT_BBC_Map!B (BBC Name) sourced from BBC_Master!A
    dv = DataValidation(type="list", formula1=f"=BBC_Master!$A$3:$A${bbc_last}", allow_blank=True)
    dv.error = "Pick a BBC Name from the BBC_Master sheet."
    dv.errorTitle = "Invalid BBC Name"
    ws_olt.add_data_validation(dv)
    dv.add(f"B3:B{max(olt_last, 2000)}")

    wbo.save(OUT)
    print("Saved workbook (with master sheets):", OUT)


    # ---------------------------------------------------------------------------
    # 7) HTML export (mirrors ExportHTMLDashboard in the VBA module)
    # ---------------------------------------------------------------------------
    rows_html = ""
    labels_js, target_js, net_js, pct_js = "", "", "", ""
    for d in detail_rows:
        net_class = "neg" if d["net"] < 0 else "pos"
        rows_html += (f"<tr><td>{d['sno']}</td><td>{d['mgr']}</td><td>{d['name']}</td><td>{d['area']}</td>"
                      f"<td>{d['olt']}</td><td>{d['target']}</td><td>{d['today']}</td><td>{d['cum']}</td>"
                      f"<td>{d['pct']*100:.2f}%</td><td>{d['clsvo']}</td><td>{d['clsnp']}</td><td>{d['disc']}</td>"
                      f"<td class='{net_class}'>{d['net']}</td></tr>\n")
        labels_js += f"'{d['name']}',"
        target_js += f"{d['target']},"
        net_js += f"{d['net']},"
        pct_js += f"{d['pct']*100:.2f},"

    tot_today = tot["today"]; tot_cum = tot["cum"]
    if has_unmapped:
        rows_html += (f"<tr style='font-style:italic;background:#fff3cd;'><td>*</td><td>(unassigned)</td>"
                      f"<td>UNMAPPED / OTHER</td><td>-</td><td>0</td><td>0</td><td>{unm_npc}</td><td>{unm_cum}</td>"
                      f"<td>-</td><td>{unm_clsvo}</td><td>{unm_clsnp}</td><td>{unm_disc}</td><td>{unm_net}</td></tr>\n")
        tot_today += unm_npc
        tot_cum += unm_cum
        tot["clsvo"] += unm_clsvo
        tot["clsnp"] += unm_clsnp

    tot_net = tot_cum - (tot["clsvo"] + tot["clsnp"])
    tot_pct = (tot_cum / tot["target"] * 100) if tot["target"] else 0

    html = f"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
    <title>FTTH Warangal Dashboard</title>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.0/chart.umd.min.js"></script>
    <style>
    body{{font-family:Segoe UI,Arial,sans-serif;background:#f2f4f8;margin:0;color:#222;}}
    .banner{{background:linear-gradient(90deg,#373c6f,#1f4eba);color:#fff;padding:22px 30px;display:flex;align-items:center;gap:18px;}}
    .logo-img{{height:58px;width:auto;object-fit:contain;display:block;}}
    .banner h1{{margin:0;font-size:26px;letter-spacing:1px;}}
    .banner .sub{{opacity:.85;font-size:13px;margin-top:4px;}}
    .kpis{{display:flex;gap:16px;padding:20px 30px 0 30px;flex-wrap:wrap;}}
    .kpi{{background:#fff;border-radius:10px;box-shadow:0 2px 6px rgba(0,0,0,.08);padding:16px 22px;min-width:150px;}}
    .kpi .v{{font-size:26px;font-weight:700;color:#1f4eba;}}
    .kpi .l{{font-size:12px;color:#666;text-transform:uppercase;letter-spacing:.5px;}}
    .kpi.neg .v{{color:#c80000;}}
    .wrap{{padding:24px 30px;}}
    table{{border-collapse:collapse;width:100%;background:#fff;box-shadow:0 2px 6px rgba(0,0,0,.06);}}
    th,td{{border:1px solid #ddd;padding:8px 10px;text-align:center;font-size:13px;}}
    th{{background:#373c6f;color:#fff;}}
    tr:nth-child(even){{background:#f7f9fc;}}
    td.neg{{color:#c80000;font-weight:700;background:#fde8e8;}}
    td.pos{{color:#0a8a0a;font-weight:700;background:#e8f8e8;}}
    .charts{{display:flex;gap:20px;flex-wrap:wrap;margin-top:24px;}}
    .chartbox{{background:#fff;border-radius:10px;box-shadow:0 2px 6px rgba(0,0,0,.08);padding:16px;flex:1;min-width:420px;}}
    .note{{font-size:12px;color:#888;margin-top:10px;font-style:italic;}}
    button.dl{{margin-top:16px;background:#1f4eba;color:#fff;border:0;padding:10px 18px;border-radius:6px;cursor:pointer;font-size:14px;}}
    button.dl:hover{{background:#173a8f;}}
    </style></head><body>
    <div class="banner"><img class="logo-img" src="data:image/png;base64,{BSNL_LOGO_B64}" alt="BSNL Logo"><div><h1>FTTH WARANGAL DASHBOARD</h1>
    <div class="sub">BBM Wise Provisioning Report of WGL OA as on {TODAY:%d-%b-%Y} &middot; source export last updated 16-AUG-2026 05:36:54</div></div></div>
    <div class="kpis">
    <div class="kpi"><div class="v">{tot['target']}</div><div class="l">Monthly Target</div></div>
    <div class="kpi"><div class="v">{tot_today}</div><div class="l">Today's Provision</div></div>
    <div class="kpi"><div class="v">{tot_cum}</div><div class="l">Cumulative Achievement</div></div>
    <div class="kpi"><div class="v">{tot_pct:.2f}%</div><div class="l">% Achievement</div></div>
    <div class="kpi"><div class="v">{tot['clsvo']+tot['clsnp']}</div><div class="l">Total Disconnections</div></div>
    <div class="kpi {'neg' if tot_net<0 else ''}"><div class="v">{tot_net}</div><div class="l">NET</div></div>
    </div>
    <div class="wrap">
    <table><thead><tr><th>S.No</th><th>AGM/Manager(MT)</th><th>BBM Name</th><th>Area</th>
    <th>OLTEs Mapped</th><th>Monthly Target</th><th>Daily Provision</th><th>Cumulative Achievement</th>
    <th>% Achievement</th><th>CLSVO</th><th>CLSNP</th><th>Total Disc.</th><th>NET</th></tr></thead><tbody>
    {rows_html}</tbody></table>
    <div class="note">* UNMAPPED / OTHER = orders whose BBC Name (rebuilt from OLT IP, or the source's own value where the IP
    wasn't on the master map) does not match one of the 11 canonical BBC entries in the master table -
    see the notes in the accompanying workbook.</div>
    <button class="dl" onclick="downloadCsv()">Download table as CSV</button>
    <div class="charts">
    <div class="chartbox"><canvas id="targetNetChart" height="260"></canvas></div>
    <div class="chartbox"><canvas id="pctChart" height="260"></canvas></div>
    </div></div>
    <script>
    const labels=[{labels_js}];const targets=[{target_js}];const nets=[{net_js}];const pcts=[{pct_js}];
    new Chart(document.getElementById('targetNetChart'),{{type:'bar',data:{{labels:labels,datasets:[
    {{label:'Monthly Target',data:targets,backgroundColor:'#1f4eba'}},
    {{label:'NET',data:nets,backgroundColor:'#c80000'}}]}},
    options:{{responsive:true,plugins:{{title:{{display:true,text:'BBC Wise: Monthly Target vs NET'}}}}}}}});
    new Chart(document.getElementById('pctChart'),{{type:'line',data:{{labels:labels,datasets:[
    {{label:'% of Achievement',data:pcts,borderColor:'#c80000',backgroundColor:'rgba(200,0,0,.15)',fill:true,tension:.3}}]}},
    options:{{responsive:true,plugins:{{title:{{display:true,text:'BBC Wise: % of Achievement'}}}}}}}});
    function downloadCsv(){{const rows=[...document.querySelectorAll('table tr')].map(r=>[...r.children]
    .map(c=>'"'+c.innerText.replace(/"/g,'""')+'"').join(','));const blob=new Blob([rows.join('\\n')],
    {{type:'text/csv'}});const a=document.createElement('a');a.href=URL.createObjectURL(blob);
    a.download='FTTH_Warangal_Dashboard.csv';a.click();}}
    </script></body></html>"""

    with open(HTML_OUT, "w", encoding="utf-8") as f:
        f.write(html)
    print("Saved HTML:", HTML_OUT)

    # ---------------------------------------------------------------------------
    # 8) "Dashboard" sheet - executive KPI-card + 2x3 chart-grid layout, styled
    #    after the reference template (banner + circular KPI icons + chart grid),
    #    populated with real FTTH Warangal data.
    # ---------------------------------------------------------------------------
    from openpyxl.chart import PieChart, Reference as ChartRef
    from openpyxl.chart.label import DataLabelList
    from openpyxl.drawing.line import LineProperties

    wbo2 = load_workbook(OUT)
    if "Dashboard" in wbo2.sheetnames:
        del wbo2["Dashboard"]
    wsd = wbo2.create_sheet("Dashboard", 0)
    wsd.sheet_view.showGridLines = False

    # ---- title banner ----
    wsd.merge_cells("A1:R3")
    wsd["A1"] = "FTTH WARANGAL DASHBOARD"
    wsd["A1"].font = Font(size=24, bold=True, color=WHITE, name="Arial")
    wsd["A1"].fill = PatternFill("solid", fgColor=NAVY)
    wsd["A1"].alignment = Alignment(horizontal="center", vertical="center")
    for rr in (1, 2, 3):
        wsd.row_dimensions[rr].height = 22
    wsd.merge_cells("S1:X3")
    wsd["S1"] = f"Data as of:\n{TODAY:%d %b, %Y}"
    wsd["S1"].font = Font(bold=True, color=NAVY, name="Arial", size=10)
    wsd["S1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wsd["S1"].fill = PatternFill("solid", fgColor=WHITE)

    # Embedded BSNL logo - no external image file required
    logo = XLImage(BytesIO(base64.b64decode(BSNL_LOGO_B64)))
    logo.width = 130
    logo.height = 59
    wsd.add_image(logo, "A1")


    # ---- KPI cards row (6 cards, matching the reference's 6-card band) ----
    kpi_defs = [
        ("TARGET", f"{tot['target']:,}", "1F4EBA"),
        ("ACHIEVED", f"{tot_cum:,}", "2E7D32"),
        ("DISCONNECTIONS", f"{tot['clsvo']+tot['clsnp']:,}", "C80000"),
        ("TODAY", f"{tot_today:,}", "6A1B9A"),
        ("% ACHIEVEMENT", f"{tot_pct:.1f}%", "EF6C00"),
        ("NET", f"{tot_net:+,}", "2E7D32" if tot_net >= 0 else "C80000"),
    ]
    kpi_col_span = 4  # each card spans 4 columns across 24 total
    row_kpi = 5
    wsd.row_dimensions[row_kpi].height = 34
    wsd.row_dimensions[row_kpi + 1].height = 16
    for i, (label, value, color) in enumerate(kpi_defs):
        c0 = i * kpi_col_span + 1
        c1 = c0 + kpi_col_span - 1
        start = get_column_letter(c0); end = get_column_letter(c1)
        wsd.merge_cells(f"{start}{row_kpi}:{end}{row_kpi}")
        cell = wsd[f"{start}{row_kpi}"]
        cell.value = value
        cell.font = Font(size=16, bold=True, color=color, name="Arial")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor="F2F4F8")
        wsd.merge_cells(f"{start}{row_kpi+1}:{end}{row_kpi+1}")
        lbl = wsd[f"{start}{row_kpi+1}"]
        lbl.value = label
        lbl.font = Font(size=8, bold=True, color="666666", name="Arial")
        lbl.alignment = Alignment(horizontal="center")

    # ---- chart grid: 2 rows x 3 cols ----
    CHART_W, CHART_H = 16, 9

    # 1) BBC-wise Cumulative Achievement (bar)
    c1 = BarChart(); c1.type = "col"; c1.title = "Provisioning by BBC (Cumulative)"
    c1.y_axis.title = "Orders"
    cats = ChartRef(ws, min_col=3, min_row=FIRST + 1, max_row=last_data_row)
    vals = ChartRef(ws, min_col=9, min_row=HDR3, max_row=last_data_row)
    c1.add_data(vals, titles_from_data=True); c1.set_categories(cats)
    c1.series[0].graphicalProperties.solidFill = "1F4EBA"
    c1.width, c1.height = CHART_W, CHART_H
    c1.legend = None
    wsd.add_chart(c1, "A8")

    # 2) Connection Type split (doughnut-style pie), sourced from real totals
    c2 = PieChart(); c2.title = "Connection Type Split"
    wsd["Z1"] = "Type"; wsd["AA1"] = "Count"
    type_order = ["NPC", "RECONNECTION", "VOLUNTAORY DISCONNECTION (CLSVO)", "DUE TO NON PAYMENT (CLSNP)"]
    type_short = {"NPC": "NPC", "RECONNECTION": "RECONNECTION",
                  "VOLUNTAORY DISCONNECTION (CLSVO)": "CLSVO", "DUE TO NON PAYMENT (CLSNP)": "CLSNP"}
    for ti, tname in enumerate(type_order, start=2):
        wsd.cell(row=ti, column=26, value=type_short[tname])   # col Z
        wsd.cell(row=ti, column=27, value=type_counts.get(tname, 0))  # col AA
    pie_labels = ChartRef(wsd, min_col=26, min_row=2, max_row=5)
    pie_data = ChartRef(wsd, min_col=27, min_row=1, max_row=5)
    c2.add_data(pie_data, titles_from_data=True)
    c2.set_categories(pie_labels)
    c2.dataLabels = DataLabelList(); c2.dataLabels.showPercent = True
    c2.width, c2.height = CHART_W, CHART_H
    wsd.add_chart(c2, "H8")

    # 3) Manager-wise NET (bar) - aggregate detail_rows by manager
    mgr_net = defaultdict(int)
    mgr_target = defaultdict(int)
    for d in detail_rows:
        mgr_net[d["mgr"]] += d["net"]
        mgr_target[d["mgr"]] += d["target"]
    wsd["Z8"] = "Manager"; wsd["AA8"] = "NET"
    mrow = 9
    for m in mgr_net:
        wsd.cell(row=mrow, column=26, value=m)
        wsd.cell(row=mrow, column=27, value=mgr_net[m])
        mrow += 1
    mgr_last = mrow - 1
    c3 = BarChart(); c3.type = "col"; c3.title = "Manager Wise NET"
    c3.add_data(ChartRef(wsd, min_col=27, min_row=8, max_row=mgr_last), titles_from_data=True)
    c3.set_categories(ChartRef(wsd, min_col=26, min_row=9, max_row=mgr_last))
    c3.series[0].graphicalProperties.solidFill = "2E7D32"
    c3.width, c3.height = CHART_W, CHART_H
    c3.legend = None
    wsd.add_chart(c3, "O8")

    # 4) Daily Provisioning Trend (line) - NPC count per completion date
    daily_npc = defaultdict(int)
    for rr in prepared:
        if rr[-2] == "NPC":
            d = rr[-3]  # "DD-Mon-YYYY" string
            daily_npc[d] += 1

    def sort_key(s):
        return datetime.datetime.strptime(s, "%d-%b-%Y")
    daily_sorted = sorted(daily_npc.items(), key=lambda kv: sort_key(kv[0]))
    wsd["Z16"] = "Date"; wsd["AA16"] = "NPC"
    drow = 17
    for dstr, cnt in daily_sorted:
        wsd.cell(row=drow, column=26, value=dstr)
        wsd.cell(row=drow, column=27, value=cnt)
        drow += 1
    daily_last = drow - 1
    c4 = LineChart(); c4.title = "Daily Provisioning Trend (NPC)"
    c4.add_data(ChartRef(wsd, min_col=27, min_row=16, max_row=daily_last), titles_from_data=True)
    c4.set_categories(ChartRef(wsd, min_col=26, min_row=17, max_row=daily_last))
    c4.series[0].graphicalProperties.line.solidFill = "C80000"
    c4.series[0].marker.symbol = "circle"
    c4.series[0].smooth = False
    c4.width, c4.height = CHART_W, CHART_H
    c4.legend = None
    wsd.add_chart(c4, "A26")

    # 5) Target vs Cumulative Achievement (grouped bar) per BBC
    c5 = BarChart(); c5.type = "col"; c5.title = "Target vs Cumulative Achievement"
    c5.add_data(ChartRef(ws, min_col=7, min_row=HDR3, max_row=last_data_row), titles_from_data=True)
    c5.add_data(ChartRef(ws, min_col=9, min_row=HDR3, max_row=last_data_row), titles_from_data=True)
    c5.set_categories(cats)
    c5.series[0].graphicalProperties.solidFill = "9E9E9E"
    c5.series[1].graphicalProperties.solidFill = "1F4EBA"
    c5.width, c5.height = CHART_W, CHART_H
    wsd.add_chart(c5, "H26")

    # 6) Disconnections by BBC (horizontal bar) - CLSVO + CLSNP stacked
    c6 = BarChart(); c6.type = "bar"; c6.title = "Disconnections by BBC"
    c6.add_data(ChartRef(ws, min_col=11, min_row=HDR3, max_row=last_data_row), titles_from_data=True)
    c6.add_data(ChartRef(ws, min_col=12, min_row=HDR3, max_row=last_data_row), titles_from_data=True)
    c6.set_categories(cats)
    c6.grouping = "stacked"; c6.overlap = 100
    c6.series[0].graphicalProperties.solidFill = "F9A825"
    c6.series[1].graphicalProperties.solidFill = "C80000"
    c6.width, c6.height = CHART_W, CHART_H
    wsd.add_chart(c6, "O26")

    # tuck the helper chart-data columns (Z:AA) out of the way
    wsd.column_dimensions["Z"].width = 1
    for col in "ABCDEFGHIJKLMNOPQRSTUVWXY":
        if col not in ("Z",):
            wsd.column_dimensions[col].width = 4.2

    wbo2.move_sheet("Dashboard", offset=-len(wbo2.sheetnames))
    wbo2.save(OUT)
    print("Saved workbook with Dashboard sheet:", OUT)
    return OUT, HTML_OUT, f"Input: {SRC.name}\nExcel output: {OUT.name}\nHTML output: {HTML_OUT.name}\nStatus: Completed successfully."
